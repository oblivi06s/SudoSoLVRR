#!/usr/bin/env python3
"""
Ablation Study Script for Sudoku Solver

Runs experiments with varying parameter values and outputs results to Excel format.

Example:
    python scripts/AblationScript.py --parameter ants --values 4 5 10 15 20 --alg 4 --output results/ablation_ants.xlsx
"""
from __future__ import annotations

import argparse
import os
import re
import sys
import statistics
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from subprocess import CompletedProcess, run, PIPE
from typing import Dict, List, Optional, Sequence, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# Import helper functions from run_general.py
# We'll define them here to avoid circular imports
def find_repo_root() -> Path:
    """Find the repository root directory."""
    return Path(__file__).resolve().parents[1]


def default_solver_candidates() -> Sequence[str]:
    """Get default solver executable candidates."""
    if os.name == "nt":
        return (
            "sudoku_ants.exe",
            "sudoku_ants",
            os.path.join("vs2017", "x64", "Release", "sudoku_ants.exe"),
            os.path.join("vs2017", "Release", "sudoku_ants.exe"),
        )
    return (
        "./sudoku_ants",
        "sudoku_ants",
        os.path.join("vs2017", "x64", "Release", "sudoku_ants"),
        os.path.join("vs2017", "Release", "sudoku_ants"),
    )


def resolve_solver_path(user_value: Optional[str]) -> Path:
    """Resolve the path to the solver executable."""
    repo_root = find_repo_root()
    if user_value:
        solver_path = Path(user_value)
        if not solver_path.is_absolute():
            solver_path = (repo_root / solver_path).resolve()
        if solver_path.exists():
            return solver_path
        raise FileNotFoundError(f"Solver binary not found at '{solver_path}'.")

    for candidate in default_solver_candidates():
        candidate_path = (repo_root / candidate).resolve()
        if candidate_path.exists() and candidate_path.is_file():
            return candidate_path

    raise FileNotFoundError(
        "Solver binary not found. Build it first (e.g. run `make -f Makefile`)."
    )


def natural_sort_key(path: Path) -> list:
    """Generate a sort key that handles numeric parts naturally."""
    parts = []
    for part in re.split(r'(\d+)', path.name):
        if part.isdigit():
            parts.append(int(part))
        else:
            parts.append(part)
    return parts


def iter_instance_files(instances_root: Path) -> Sequence[Path]:
    """Iterate over instance files in the given directory."""
    if not instances_root.exists():
        raise FileNotFoundError(f"Instances folder not found: '{instances_root}'.")
    all_files = list(instances_root.glob("*.txt"))
    all_files.extend([f for f in instances_root.glob("*") if f.is_file() and f.suffix == ""])
    return sorted(all_files, key=natural_sort_key)


@dataclass(frozen=True)
class InstanceMetadata:
    """Metadata about a Sudoku instance."""
    path: Path
    size_label: str
    fixed_percentage: Optional[int]
    instance_id: Optional[int]

    @property
    def relative_path(self) -> str:
        repo_root = find_repo_root()
        return str(self.path.relative_to(repo_root))


def parse_metadata(path: Path) -> InstanceMetadata:
    """Parse metadata from instance file path."""
    name = path.stem
    match = re.match(r"ins(?P<size>[0-9x]+)_(?P<fixed>\d+)_(?P<idx>\d+)", name)
    size = None
    fixed = None
    idx = None
    if match:
        size = match.group("size")
        fixed = int(match.group("fixed"))
        idx = int(match.group("idx"))
    else:
        size = name
    return InstanceMetadata(path=path, size_label=size or "unknown", fixed_percentage=fixed, instance_id=idx)


def format_instance_argument(instance_path: Path, repo_root: Path) -> str:
    """Format instance path for command-line argument."""
    try:
        rel_path = instance_path.relative_to(repo_root)
    except ValueError:
        path_str = str(instance_path)
    else:
        prefixed = Path(".") / rel_path
        path_str = str(prefixed)
    
    if ' ' in path_str:
        return f'"{path_str}"'
    return path_str


def build_solver_command(
    solver_path: Path,
    instance_path: Path,
    repo_root: Path,
    algorithm: int,
    timeout: float,
    parameter_name: str,
    parameter_value: any,
    constant_params: Dict[str, any],
) -> List[str]:
    """Build solver command with given parameters."""
    file_arg = format_instance_argument(instance_path, repo_root)
    cmd: List[str] = [str(solver_path), "--file", file_arg, "--alg", str(algorithm), "--timeout", str(timeout)]

    # Add all constant parameters
    param_mapping = {
        "ants": "--ants",
        "q0": "--q0",
        "rho": "--rho",
        "evap": "--evap",
        "threads": "--threads",
        "numacs": "--numacs",
        "numcolonies": "--numcolonies",
        "subcolonies": "--subcolonies",
        "convthreshold": "--convthreshold",
        "entropythreshold": "--entropythreshold",
        "comm-early-interval": "--comm-early-interval",
        "comm-late-interval": "--comm-late-interval",
        "comm-threshold": "--comm-threshold",
    }

    # Add constant parameters
    for param_name, param_value in constant_params.items():
        if param_name in param_mapping and param_value is not None:
            cmd.extend((param_mapping[param_name], str(param_value)))

    # Override with the varying parameter
    if parameter_name in param_mapping:
        cmd.extend((param_mapping[parameter_name], str(parameter_value)))

    # Always add verbose for algorithms 0, 2, 3, 4
    if algorithm in [0, 2, 3, 4]:
        cmd.append("--verbose")

    return cmd


def run_solver(cmd: Sequence[str], cwd: Path, timeout: Optional[float]) -> CompletedProcess:
    """Run the solver executable."""
    return run(
        list(cmd),
        cwd=str(cwd),
        capture_output=True,
        text=True,
        timeout=timeout,
    )


def parse_solver_output(stdout: str, stderr: str) -> Tuple[Optional[bool], Optional[float], Optional[int]]:
    """Parse solver output to extract success, time, and iterations."""
    stdout_lines = [line.strip() for line in stdout.splitlines() if line.strip()]
    stderr_lines = [line.strip() for line in stderr.splitlines() if line.strip()]

    success: Optional[bool] = None
    solve_time: Optional[float] = None
    iterations: Optional[int] = None

    all_lines = stdout_lines + stderr_lines

    for line in all_lines:
        if line in {"0", "1"} and success is None:
            success = (line == "0")
            continue

        solved_match = re.search(r"[Ss]olved in ([0-9]*\.?[0-9]+)", line)
        if solved_match:
            solve_time = float(solved_match.group(1))
            success = True
            continue

        failed_match = re.search(r"[Ff]ailed (?:in time |to solve in )([0-9]*\.?[0-9]+)", line)
        if failed_match:
            solve_time = float(failed_match.group(1))
            success = False
            continue

        iter_match = re.search(r"[Ii]terations:\s*([0-9]+)", line)
        if iter_match:
            iterations = int(iter_match.group(1))
            continue

        total_time_match = re.search(r"Total Solve Time:\s*([0-9]*\.?[0-9]+)\s+s", line)
        if total_time_match:
            solve_time = float(total_time_match.group(1))
            continue

    # Fallback: check stdout for time if not found yet
    for line in stdout_lines:
        if solve_time is None:
            try:
                solve_time = float(line)
            except ValueError:
                pass

    if success is None:
        for line in stderr_lines:
            if "could not open file" in line.lower():
                success = False
                break

    if solve_time is not None:
        solve_time = round(solve_time, 5)

    return success, solve_time, iterations


def calculate_statistics(results_list: List[Tuple[bool, float, int]]) -> Dict[str, float]:
    """Calculate statistics from a list of results."""
    if not results_list:
        return {
            "success_rate": 0.0,
            "avg_time": 0.0,
            "std_dev": 0.0,
            "avg_iteration": 0.0,
        }

    successes = sum(1 for success, _, _ in results_list if success)
    total = len(results_list)
    success_rate = (successes / total * 100.0) if total > 0 else 0.0

    # Only include successful runs in time and iteration statistics
    successful_times = [time for success, time, _ in results_list if success and time is not None]
    successful_iterations = [iter for success, _, iter in results_list if success and iter is not None]

    avg_time = statistics.mean(successful_times) if successful_times else 0.0
    std_dev = statistics.stdev(successful_times) if len(successful_times) > 1 else 0.0
    avg_iteration = statistics.mean(successful_iterations) if successful_iterations else 0.0

    return {
        "success_rate": round(success_rate, 2),
        "avg_time": round(avg_time, 5),
        "std_dev": round(std_dev, 5),
        "avg_iteration": round(avg_iteration, 2),
    }


def write_excel_output(
    output_path: Path,
    constant_params: Dict[str, any],
    results: Dict[str, Dict[str, Dict[str, float]]],
    parameter_name: str,
    algorithm: int,
) -> None:
    """Write results to Excel file with proper formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ablation Results"

    # Define styles
    header_font = Font(bold=True, size=12)
    param_label_font = Font(bold=True, size=11)
    value_font = Font(size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    row = 1

    # Write constant parameters section
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws[f'A{row}']
    cell.value = "Constant Parameters"
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = header_fill
    cell.border = border
    row += 1

    # Write parameter headers
    ws[f'A{row}'] = "Parameter"
    ws[f'B{row}'] = "Value"
    ws[f'A{row}'].font = param_label_font
    ws[f'B{row}'].font = param_label_font
    ws[f'A{row}'].border = border
    ws[f'B{row}'].border = border
    row += 1

    # Write constant parameters
    param_display_names = {
        "q0": "q0",
        "rho": "rho",
        "evap": "evap",
        "ants": "ants",
        "threads": "threads",
        "numacs": "numacs",
        "numcolonies": "numcolonies",
        "timeout": "timeout",
        "convthreshold": "convthreshold",
        "entropythreshold": "entropythreshold",
        "comm-early-interval": "comm-early-interval",
        "comm-late-interval": "comm-late-interval",
        "comm-threshold": "comm-threshold",
    }

    # Always show algorithm
    ws[f'A{row}'] = "algorithm"
    ws[f'B{row}'] = algorithm
    ws[f'A{row}'].font = value_font
    ws[f'B{row}'].font = value_font
    ws[f'A{row}'].border = border
    ws[f'B{row}'].border = border
    row += 1

    # Show other constant parameters (excluding the varying parameter)
    for param_key in sorted(constant_params.keys()):
        if param_key != parameter_name and constant_params[param_key] is not None:
            display_name = param_display_names.get(param_key, param_key)
            ws[f'A{row}'] = display_name
            ws[f'B{row}'] = constant_params[param_key]
            ws[f'A{row}'].font = value_font
            ws[f'B{row}'].font = value_font
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1

    # Empty row
    row += 1

    # Write results header
    ws[f'A{row}'] = "Parameter Value"
    ws[f'B{row}'] = "Instance"
    ws[f'C{row}'] = "Success Rate (%)"
    ws[f'D{row}'] = "Average Time (s)"
    ws[f'E{row}'] = "Std Dev (s)"
    ws[f'F{row}'] = "Avg Iteration"

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        cell = ws[f'{col}{row}']
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
        cell.border = border

    row += 1

    # Write results for each parameter value (sorted numerically)
    def sort_key(x: str) -> float:
        """Extract numeric value for sorting."""
        try:
            return float(x)
        except ValueError:
            # If not a number, try to extract first number
            match = re.search(r'[\d.]+', x)
            return float(match.group()) if match else 0.0
    
    for param_value_str in sorted(results.keys(), key=sort_key):
        instance_results = results[param_value_str]

        # Write parameter value label (merged cells for first column)
        ws.merge_cells(f'A{row}:A{row + len(instance_results) - 1}')
        cell = ws[f'A{row}']
        cell.value = f"{parameter_name} = {param_value_str}"
        cell.font = param_label_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

        # Write results for each instance
        for instance_name in sorted(instance_results.keys()):
            stats = instance_results[instance_name]
            ws[f'B{row}'] = instance_name
            ws[f'C{row}'] = stats["success_rate"]
            ws[f'D{row}'] = stats["avg_time"]
            ws[f'E{row}'] = stats["std_dev"]
            ws[f'F{row}'] = stats["avg_iteration"]

            # Apply formatting
            for col in ['B', 'C', 'D', 'E', 'F']:
                cell = ws[f'{col}{row}']
                cell.font = value_font
                cell.border = border
                if col == 'C':  # Success rate percentage
                    cell.number_format = '0.00'
                elif col in ['D', 'E']:  # Time values
                    cell.number_format = '0.00000'
                elif col == 'F':  # Iterations
                    cell.number_format = '0.00'

            row += 1

        # Empty row between parameter groups
        row += 1

    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    # Save workbook
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"\nResults saved to: {output_path}")


def run_ablation_experiment(
    parameter_name: str,
    parameter_values: List[any],
    constant_params: Dict[str, any],
    instances: List[InstanceMetadata],
    solver_path: Path,
    repo_root: Path,
    timeout: float,
    algorithm: int,
    runs_per_instance: int = 1,
    verbose: bool = True,
) -> Dict[str, Dict[str, Dict[str, float]]]:
    """Run ablation experiment and return results."""
    results: Dict[str, Dict[str, List[Tuple[bool, float, int]]]] = defaultdict(lambda: defaultdict(list))

    total_experiments = len(parameter_values) * len(instances) * runs_per_instance
    current_experiment = 0

    for param_value in parameter_values:
        param_value_str = str(param_value)
        if verbose:
            print(f"\n{'='*60}")
            print(f"Testing {parameter_name} = {param_value_str}")
            print(f"{'='*60}")
        
        counter = 1
        for instance_meta in instances:
            # Determine number of runs (100 for logic-solvable, runs_per_instance for general)
            is_logic_solvable = instance_meta.fixed_percentage is None
            num_runs = 100 if is_logic_solvable else runs_per_instance

            if verbose:
                print(f"\nProcessing: {instance_meta.relative_path} ({instance_meta.size_label}) ({counter})")

            for run_num in range(1, num_runs + 1):
                current_experiment += 1
                if verbose and num_runs > 1:
                    print(f"  Run {run_num}/{num_runs}...", end=' ', flush=True)

                cmd = build_solver_command(
                    solver_path,
                    instance_meta.path,
                    repo_root,
                    algorithm,
                    timeout,
                    parameter_name,
                    param_value,
                    constant_params,
                )

                try:
                    result = run_solver(cmd, repo_root, timeout=None)
                    success, solve_time, iterations = parse_solver_output(result.stdout, result.stderr or "")

                    if success is None:
                        success = False
                    if solve_time is None:
                        solve_time = timeout if not success else 0.0

                    results[param_value_str][instance_meta.size_label].append((success, solve_time, iterations))

                    if verbose and num_runs > 1:
                        status = "OK" if success else "FAIL"
                        print(f"{status}", end='', flush=True)
                        if run_num < num_runs:
                            print(" ", end='', flush=True)

                except Exception as e:
                    if verbose:
                        print(f"ERROR: {e}", file=sys.stderr)
                    results[param_value_str][instance_meta.size_label].append((False, timeout, 0))

            if verbose and num_runs > 1:
                print()  # New line after all runs

            # Calculate and display statistics for this instance
            if verbose:
                stats = calculate_statistics(results[param_value_str][instance_meta.size_label])
                print(f"  {instance_meta.size_label}: Success Rate={stats['success_rate']:.2f}%, "
                      f"Avg Time={stats['avg_time']:.5f}s, Avg Iter={stats['avg_iteration']:.2f}")
            
            counter += 1

    # Convert to final format with calculated statistics
    final_results: Dict[str, Dict[str, Dict[str, float]]] = {}
    for param_value_str, instance_dict in results.items():
        final_results[param_value_str] = {}
        for instance_name, result_list in instance_dict.items():
            final_results[param_value_str][instance_name] = calculate_statistics(result_list)

    return final_results


def main() -> int:
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Run ablation studies on Sudoku solver algorithms",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Test different ant counts
  python scripts/AblationScript.py --parameter ants --values 4 5 10 15 20 --alg 4 --output results/ablation_ants.xlsx

  # Test different q0 values
  python scripts/AblationScript.py --parameter q0 --values 0.0 0.3 0.5 0.7 0.9 --alg 4 --output results/ablation_q0.xlsx
        """
    )

    parser.add_argument("--parameter", required=True, help="Parameter to vary (e.g., 'ants', 'q0', 'rho')")
    parser.add_argument("--values", required=True, nargs="+", help="Values to test for the parameter")
    parser.add_argument("--alg", type=int, default=4, help="Algorithm to use (0-4, default: 4)")
    parser.add_argument("--instances-root", default=None, help="Folder containing instances (default: instances/general)")
    parser.add_argument("--output", required=True, help="Output Excel file path")
    parser.add_argument("--solver", default=None, help="Path to solver executable (default: auto-detect)")
    parser.add_argument("--timeout", type=float, default=120.0, help="Timeout per puzzle in seconds (default: 120)")

    # Constant parameters
    parser.add_argument("--ants", type=int, default=None, help="Number of ants (constant)")
    parser.add_argument("--q0", type=float, default=None, help="q0 parameter (constant)")
    parser.add_argument("--rho", type=float, default=None, help="rho parameter (constant)")
    parser.add_argument("--evap", type=float, default=None, help="evap parameter (constant)")
    parser.add_argument("--threads", type=int, default=None, help="Number of threads (constant)")
    parser.add_argument("--numacs", type=int, default=None, help="Number of ACS colonies (constant)")
    parser.add_argument("--numcolonies", type=int, default=None, help="Total number of colonies (constant)")
    parser.add_argument("--subcolonies", type=int, default=None, help="Number of sub-colonies (constant)")
    parser.add_argument("--convthreshold", type=float, default=None, help="Convergence threshold (constant)")
    parser.add_argument("--entropythreshold", type=float, default=None, help="Entropy threshold (constant)")
    parser.add_argument("--comm-early-interval", type=int, default=None, help="Early communication interval (constant)")
    parser.add_argument("--comm-late-interval", type=int, default=None, help="Late communication interval (constant)")
    parser.add_argument("--comm-threshold", type=int, default=None, help="Communication threshold (constant)")

    parser.add_argument("--runs-per-instance", type=int, default=1, help="Number of runs per general instance (default: 1; logic-solvable instances always use 100 runs)")
    parser.add_argument("--verbose", action="store_true", default=True, help="Print progress (default: True)")
    parser.add_argument("--no-verbose", dest="verbose", action="store_false", help="Disable progress output")

    args = parser.parse_args()

    # Validate parameter name
    valid_parameters = {
        "ants", "q0", "rho", "evap", "threads", "numacs", "numcolonies",
        "subcolonies", "convthreshold", "entropythreshold",
        "comm-early-interval", "comm-late-interval", "comm-threshold"
    }
    if args.parameter not in valid_parameters:
        print(f"Error: Invalid parameter '{args.parameter}'. Valid parameters: {', '.join(sorted(valid_parameters))}", file=sys.stderr)
        return 1

    # Parse parameter values (handle both int and float)
    try:
        if args.parameter in ["ants", "threads", "numacs", "numcolonies", "subcolonies", "comm-early-interval", "comm-late-interval", "comm-threshold"]:
            parameter_values = [int(v) for v in args.values]
        else:
            parameter_values = [float(v) for v in args.values]
    except ValueError as e:
        print(f"Error: Invalid parameter values: {e}", file=sys.stderr)
        return 1

    repo_root = find_repo_root()

    # Resolve solver path
    try:
        solver_path = resolve_solver_path(args.solver)
    except FileNotFoundError as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1

    # Resolve instances root
    if args.instances_root:
        instances_root = (repo_root / args.instances_root).resolve()
    else:
        instances_root = repo_root / "instances" / "general"

    if not instances_root.exists():
        print(f"Error: Instances folder not found: {instances_root}", file=sys.stderr)
        return 1

    # Get instance files
    try:
        instance_files = list(iter_instance_files(instances_root))
    except FileNotFoundError as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1

    if not instance_files:
        print(f"Error: No instance files found in '{instances_root}'", file=sys.stderr)
        return 1

    instances = [parse_metadata(path) for path in instance_files]

    # Build constant parameters dictionary
    constant_params = {
        "ants": args.ants,
        "q0": args.q0,
        "rho": args.rho,
        "evap": args.evap,
        "threads": args.threads,
        "numacs": args.numacs,
        "numcolonies": args.numcolonies,
        "subcolonies": args.subcolonies,
        "convthreshold": args.convthreshold,
        "entropythreshold": args.entropythreshold,
        "comm-early-interval": getattr(args, "comm_early_interval", None),
        "comm-late-interval": getattr(args, "comm_late_interval", None),
        "comm-threshold": getattr(args, "comm_threshold", None),
    }

    # Remove None values
    constant_params = {k: v for k, v in constant_params.items() if v is not None}

    # Resolve output path
    output_path = Path(args.output)
    if not output_path.is_absolute():
        output_path = (repo_root / output_path).resolve()

    # Print configuration
    if args.verbose:
        print("="*60)
        print("Ablation Study Configuration")
        print("="*60)
        print(f"Parameter to vary: {args.parameter}")
        print(f"Values to test: {parameter_values}")
        print(f"Algorithm: {args.alg}")
        print(f"Instances folder: {instances_root}")
        print(f"Number of instances: {len(instances)}")
        print(f"Output file: {output_path}")
        print(f"Constant parameters: {constant_params}")
        print("="*60)

    # Run experiments
    try:
        results = run_ablation_experiment(
            args.parameter,
            parameter_values,
            constant_params,
            instances,
            solver_path,
            repo_root,
            args.timeout,
            args.alg,
            args.runs_per_instance,
            args.verbose,
        )

        # Write Excel output
        write_excel_output(output_path, constant_params, results, args.parameter, args.alg)

        if args.verbose:
            print("\n" + "="*60)
            print("Ablation study completed successfully!")
            print("="*60)

        return 0

    except KeyboardInterrupt:
        print("\n\nInterrupted by user. Partial results may be saved.", file=sys.stderr)
        return 1
    except Exception as e:
        print(f"\nError during execution: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
